"""
Circuit Diagram Inspector
Interactive PDF inspection tool for circuit diagrams with error logging to Excel
"""

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Menu
from PIL import Image, ImageTk, ImageDraw
import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import json
import cv2
import numpy as np


class CircuitInspector:
    def __init__(self, root):
        self.root = root
        self.root.title("Circuit Diagram Inspector")
        self.root.geometry("1400x900")
        
        # Data storage
        self.pdf_document = None
        self.current_page = 0
        self.cabinet_id = ""
        self.project_name = ""
        self.sales_order_no = ""
        self.annotations = []  # List of {type: 'ok'/'error', x, y, error_text, contour}
        self.excel_file = "Emerson.xlsx"
        self.zoom_level = 1.0
        self.current_sr_no = 1
        self.current_page_image = None  # Store current page image for contour detection
        
        # Error categories with hierarchical structure
        self.error_categories = {
            'Material Shortfall': {
                'Component Missing': '{tag} missing',
                'TB Group Missing': '{tag} TB group missing',
                'Fuse Missing': '{tag} Fuse missing',
                'Label Missing': '{tag} Label missing',
                'KLM Marker Missing': '{tag} KLMA Marker Missing',
                'End stopper missing': 'End stopper near {tag} missing',
                'Short link missing': '{tag} TB Group {terminals} shortlink missing'
            },
            'Wrong Wiring': {
                'Wires Interchanges': '{tag} Wires Interchanges',
                'Color Code Wrong': '{tag} Color Code Wrong',
                'Ferrule Wrong': '{tag} Ferrule Wrong',
                'Size Wrong': '{tag} Size Wrong',
                'Wire Loose Found': '{tag} Wire Loose Found',
                'Lug not properly cut': '{tag} Lug not properly cut'
            },
            'Incomplete Wiring': {
                'All wiring Incomplete with connections pending': '{tag} All wiring Incomplete with connections pending',
                'Connections pending': '{tag} Connections pending'
            },
            'Wrong Assembly': {
                'Label Wrong installed': '{tag} Label Wrong installed',
                'Fuse Wrong installed': '{tag} Fuse Wrong installed',
                'Wire duct Wrong Installed': '{tag} Wire duct Wrong Installed',
                'Component Wrong installed': '{tag} Component Wrong installed',
                'Component not properly fixed': '{tag} Component not properly fixed',
                'End stopper loose found': '{tag} End stopper loose found'
            }
        }
        
        # Rectangle drawing state
        self.drawing = False
        self.drawing_type = None  # 'ok' or 'error'
        self.rect_start_x = None
        self.rect_start_y = None
        self.temp_rect_id = None
        
        # Annotation selection state
        self.selected_annotation = None
        
        self.setup_ui()
        self.setup_excel()
        
    def setup_ui(self):
        """Create the user interface"""
        # Top toolbar
        toolbar = tk.Frame(self.root, bg='#2c3e50', height=60)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        
        # Buttons
        btn_style = {'bg': '#3498db', 'fg': 'white', 'padx': 15, 'pady': 8, 'font': ('Arial', 10)}
        
        tk.Button(toolbar, text="📁 Load PDF", command=self.load_pdf, **btn_style).pack(side=tk.LEFT, padx=5, pady=10)
        tk.Button(toolbar, text="🆔 Set Cabinet ID", command=self.set_cabinet_id, **btn_style).pack(side=tk.LEFT, padx=5, pady=10)
        
        # Cabinet ID display
        self.cabinet_label = tk.Label(toolbar, text="Cabinet: Not Set", bg='#2c3e50', fg='white', font=('Arial', 11, 'bold'))
        self.cabinet_label.pack(side=tk.LEFT, padx=20)
        
        # Page navigation
        tk.Button(toolbar, text="◀ Prev", command=self.prev_page, bg='#95a5a6', fg='white', padx=10, pady=8).pack(side=tk.LEFT, padx=5)
        self.page_label = tk.Label(toolbar, text="Page: 0/0", bg='#2c3e50', fg='white', font=('Arial', 10))
        self.page_label.pack(side=tk.LEFT, padx=5)
        tk.Button(toolbar, text="Next ▶", command=self.next_page, bg='#95a5a6', fg='white', padx=10, pady=8).pack(side=tk.LEFT, padx=5)
        
        # Zoom controls
        tk.Button(toolbar, text="🔍+", command=self.zoom_in, bg='#27ae60', fg='white', padx=10, pady=8).pack(side=tk.LEFT, padx=(20, 2))
        tk.Button(toolbar, text="🔍-", command=self.zoom_out, bg='#27ae60', fg='white', padx=10, pady=8).pack(side=tk.LEFT, padx=2)
        
        # Save/Export buttons
        tk.Button(toolbar, text="📥 Export Annotated PDF", command=self.export_annotated_pdf, bg='#e67e22', fg='white', padx=10, pady=8).pack(side=tk.RIGHT, padx=5, pady=10)
        tk.Button(toolbar, text="📊 Open Excel", command=self.open_excel, bg='#16a085', fg='white', padx=10, pady=8).pack(side=tk.RIGHT, padx=5, pady=10)
        
        # Main canvas with scrollbars
        canvas_frame = tk.Frame(self.root)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        v_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Canvas
        self.canvas = tk.Canvas(canvas_frame, bg='#ecf0f1', 
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scrollbar.config(command=self.canvas.yview)
        h_scrollbar.config(command=self.canvas.xview)
        
        # Bind mouse events for drag drawing
        self.canvas.bind("<ButtonPress-1>", self.on_left_press)
        self.canvas.bind("<B1-Motion>", self.on_left_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_left_release)
        
        self.canvas.bind("<ButtonPress-3>", self.on_right_press)
        self.canvas.bind("<B3-Motion>", self.on_right_drag)
        self.canvas.bind("<ButtonRelease-3>", self.on_right_release)
        
        # Bind Ctrl+Click for selection
        self.canvas.bind("<Control-Button-1>", self.on_ctrl_click)
        
        # Bind Delete key for deleting selected annotation
        self.root.bind("<Delete>", self.delete_selected_annotation)
        
        # Instructions panel
        instructions = tk.Frame(self.root, bg='#34495e', height=50)
        instructions.pack(side=tk.BOTTOM, fill=tk.X)
        
        inst_text = "🖱️ Left Drag: OK (Green Circle) | Right Drag: Error (Yellow Box) | Ctrl+Click: Select | Delete: Remove"
        tk.Label(instructions, text=inst_text, bg='#34495e', fg='white', 
                font=('Arial', 10), pady=15).pack()
    
    def setup_excel(self):
        """Initialize or load Excel file"""
        from openpyxl.styles import Alignment, Border, Side, Font
        
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Merge cells for headers
            ws.merge_cells('C4:Q4')
            ws['C4'] = 'Project Name'
            ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('C6:I6')
            ws['C6'] = 'Sales Order No.'
            ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('J6:Q6')
            ws['J6'] = 'Cabinet ID:'
            ws['J6'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Column headers (row 9-10)
            ws.merge_cells('C9:C10')
            ws['C9'] = 'Sr No.'
            ws['C9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.merge_cells('D9:D10')
            ws['D9'] = 'Refference No.'
            ws['D9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.merge_cells('E9:E10')
            ws['E9'] = 'Punch / Action Point'
            ws['E9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.merge_cells('F9:F10')
            ws['F9'] = 'Category'
            ws['F9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.merge_cells('G9:H9')
            ws['G9'] = 'Checked By'
            ws['G9'].alignment = Alignment(horizontal='center', vertical='center')
            ws['G10'] = 'Name'
            ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
            ws['H10'] = 'Date'
            ws['H10'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('I9:J9')
            ws['I9'] = 'Implemented By(Production)'
            ws['I9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['I10'] = 'Name'
            ws['I10'].alignment = Alignment(horizontal='center', vertical='center')
            ws['J10'] = 'Date'
            ws['J10'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('K9:L9')
            ws['K9'] = 'Closed By'
            ws['K9'].alignment = Alignment(horizontal='center', vertical='center')
            ws['K10'] = 'Name'
            ws['K10'].alignment = Alignment(horizontal='center', vertical='center')
            ws['L10'] = 'Date'
            ws['L10'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            
            # Apply borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws['C4:Q6']:
                for cell in row:
                    cell.border = thin_border
            
            for row in ws['C9:L10']:
                for cell in row:
                    cell.border = thin_border
                    cell.font = Font(bold=True)
            
            wb.save(self.excel_file)
    
    def load_pdf(self):
        """Load a PDF file"""
        file_path = filedialog.askopenfilename(
            title="Select Circuit Diagram PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                self.pdf_document = fitz.open(file_path)
                self.current_page = 0
                self.annotations = []
                self.zoom_level = 1.0
                
                # Get next Sr No from Excel (continue from last entry)
                self.current_sr_no = self.get_next_sr_no()
                
                self.cabinet_label.config(text=f"Cabinet: {self.cabinet_id if self.cabinet_id else 'Not Set'}")
                
                self.display_page()
                messagebox.showinfo("Success", f"Loaded PDF with {len(self.pdf_document)} pages")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load PDF: {str(e)}")
    
    def get_next_sr_no(self):
        """Get the next Sr No by reading the last entry in Excel"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Find the last filled row in column C (Sr No column) starting from row 11
            last_sr_no = 0
            row_num = 11
            while ws[f'C{row_num}'].value is not None:
                try:
                    sr_value = ws[f'C{row_num}'].value
                    if sr_value and isinstance(sr_value, (int, float)):
                        last_sr_no = int(sr_value)
                    elif sr_value and str(sr_value).isdigit():
                        last_sr_no = int(sr_value)
                except:
                    pass
                row_num += 1
            
            wb.close()
            return last_sr_no + 1
        except Exception as e:
            print(f"Error reading Sr No from Excel: {e}")
            return 1  # Default to 1 if there's an error
    
    def set_cabinet_id(self):
        """Set the cabinet identifier"""
        new_id = simpledialog.askstring("Cabinet ID", "Enter Cabinet ID:", 
                                       initialvalue=self.cabinet_id)
        if new_id is not None:
            self.cabinet_id = new_id
            self.cabinet_label.config(text=f"Cabinet: {self.cabinet_id}")
            self.update_excel_headers()
    
    def update_excel_headers(self):
        """Update Excel file with project details"""
        try:
            # Load the existing workbook
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Update the top-left cell of each merged range directly.
            # This preserves the merge and cell styling while updating the text.
            
            # Project Name (Merged C4:Q4 -> Write to C4)
            if self.project_name:
                ws['C4'] = f"Project Name: {self.project_name}"
            
            # Sales Order No (Merged C6:I6 -> Write to C6)
            if self.sales_order_no:
                ws['C6'] = f"Sales Order No.: {self.sales_order_no}"
            
            # Cabinet ID (Merged J6:Q6 -> Write to J6)
            if self.cabinet_id:
                ws['J6'] = f"Cabinet ID: {self.cabinet_id}"
            
            # Save the file
            wb.save(self.excel_file)
            wb.close()
            
            print(f"✓ Updated Excel headers: {self.project_name}, {self.sales_order_no}, {self.cabinet_id}")
            
            messagebox.showinfo("Excel Updated", 
                              f"Excel file updated successfully!\n\n"
                              f"Project: {self.project_name}\n"
                              f"Sales Order: {self.sales_order_no}\n"
                              f"Cabinet ID: {self.cabinet_id}\n\n"
                              f"If Excel is open, close and reopen it to see changes.")
            
        except PermissionError:
             messagebox.showerror("Excel Error", "Please close the Excel file before updating headers.")
        except Exception as e:
            print(f"❌ Error updating Excel headers: {e}")
            messagebox.showerror("Excel Error", f"Failed to update Excel: {str(e)}")
    
    def detect_component_at_point(self, x, y, img_array):
        """Detect component contour at the clicked point - simplified with fixed size"""
        # Use a fixed reasonable highlight size instead of complex detection
        # This is more reliable for circuit diagrams with varying component types
        highlight_width = 40
        highlight_height = 40
        
        return (
            max(0, x - highlight_width),
            max(0, y - highlight_height),
            min(img_array.shape[1], x + highlight_width),
            min(img_array.shape[0], y + highlight_height)
        )
    
    def display_page(self):
        """Render and display the current PDF page"""
        if not self.pdf_document:
            return
        
        try:
            # Get page
            page = self.pdf_document[self.current_page]
            
            # Render at higher resolution for better quality
            mat = fitz.Matrix(2.0 * self.zoom_level, 2.0 * self.zoom_level)
            pix = page.get_pixmap(matrix=mat)
            
            # Convert to PIL Image
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Store current page image for component detection
            self.current_page_image = np.array(img)
            
            # Draw annotations
            draw = ImageDraw.Draw(img, 'RGBA')
            for ann in self.annotations:
                if ann['page'] == self.current_page and 'bbox' in ann:
                    # Scale bounding box with zoom
                    bbox = ann['bbox']
                    x1, y1, x2, y2 = bbox
                    
                    # Check if this annotation is selected
                    is_selected = (self.selected_annotation is ann)
                    
                    if ann['type'] == 'ok':
                        # Green circle - calculate center and radius from bbox
                        center_x = (x1 + x2) / 2
                        center_y = (y1 + y2) / 2
                        radius = max(abs(x2 - x1), abs(y2 - y1)) / 2
                        
                        # Draw circle
                        draw.ellipse(
                            [center_x - radius, center_y - radius, center_x + radius, center_y + radius],
                            fill=(0, 255, 0, 80),  # Semi-transparent green
                            outline='blue' if is_selected else 'green',
                            width=int(5 * self.zoom_level) if is_selected else int(3 * self.zoom_level)
                        )
                    else:
                        # Yellow semi-transparent highlight
                        draw.rectangle(
                            [x1, y1, x2, y2],
                            fill=(255, 255, 0, 100),  # Semi-transparent yellow
                            outline='blue' if is_selected else 'orange',
                            width=int(5 * self.zoom_level) if is_selected else int(3 * self.zoom_level)
                        )
            
            # Convert to PhotoImage
            self.photo = ImageTk.PhotoImage(img)
            
            # Update canvas
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            
            # Update page label
            self.page_label.config(text=f"Page: {self.current_page + 1}/{len(self.pdf_document)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to display page: {str(e)}")
    
    def on_left_press(self, event):
        """Start drawing green circle"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return
        
        self.drawing = True
        self.drawing_type = 'ok'
        self.rect_start_x = self.canvas.canvasx(event.x)
        self.rect_start_y = self.canvas.canvasy(event.y)
    
    def on_left_drag(self, event):
        """Update green circle while dragging"""
        if not self.drawing or self.drawing_type != 'ok':
            return
        
        # Remove previous temporary shape
        if self.temp_rect_id:
            self.canvas.delete(self.temp_rect_id)
        
        # Draw new temporary circle
        current_x = self.canvas.canvasx(event.x)
        current_y = self.canvas.canvasy(event.y)
        
        # Calculate radius from drag distance
        radius = ((current_x - self.rect_start_x)**2 + (current_y - self.rect_start_y)**2)**0.5
        
        self.temp_rect_id = self.canvas.create_oval(
            self.rect_start_x - radius, self.rect_start_y - radius,
            self.rect_start_x + radius, self.rect_start_y + radius,
            outline='green', width=3, dash=(5, 5)
        )
    
    def on_left_release(self, event):
        """Finalize green circle"""
        if not self.drawing or self.drawing_type != 'ok':
            return
        
        # Remove temporary shape
        if self.temp_rect_id:
            self.canvas.delete(self.temp_rect_id)
            self.temp_rect_id = None
        
        end_x = self.canvas.canvasx(event.x)
        end_y = self.canvas.canvasy(event.y)
        
        # Calculate radius from drag distance
        radius = ((end_x - self.rect_start_x)**2 + (end_y - self.rect_start_y)**2)**0.5
        
        # Minimum circle size
        if radius < 5:
            self.drawing = False
            return
        
        # Store as bbox for consistency (circle will be drawn from center and radius)
        center_x = self.rect_start_x
        center_y = self.rect_start_y
        bbox = (center_x - radius, center_y - radius, center_x + radius, center_y + radius)
        
        # Add OK annotation
        self.annotations.append({
            'type': 'ok',
            'page': self.current_page,
            'x': center_x / (2.0 * self.zoom_level),
            'y': center_y / (2.0 * self.zoom_level),
            'bbox': bbox,
            'timestamp': datetime.now().isoformat()
        })
        
        self.drawing = False
        self.display_page()
    
    def on_right_press(self, event):
        """Start drawing yellow rectangle"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return
        
        self.drawing = True
        self.drawing_type = 'error'
        self.rect_start_x = self.canvas.canvasx(event.x)
        self.rect_start_y = self.canvas.canvasy(event.y)
    
    def on_right_drag(self, event):
        """Update yellow rectangle while dragging"""
        if not self.drawing or self.drawing_type != 'error':
            return
        
        # Remove previous temporary rectangle
        if self.temp_rect_id:
            self.canvas.delete(self.temp_rect_id)
        
        # Draw new temporary rectangle
        current_x = self.canvas.canvasx(event.x)
        current_y = self.canvas.canvasy(event.y)
        self.temp_rect_id = self.canvas.create_rectangle(
            self.rect_start_x, self.rect_start_y, current_x, current_y,
            outline='yellow', width=3, dash=(5, 5)
        )
    
    def on_right_release(self, event):
        """Finalize yellow rectangle and show error menu"""
        if not self.drawing or self.drawing_type != 'error':
            return
        
        # Remove temporary rectangle
        if self.temp_rect_id:
            self.canvas.delete(self.temp_rect_id)
            self.temp_rect_id = None
        
        end_x = self.canvas.canvasx(event.x)
        end_y = self.canvas.canvasy(event.y)
        
        # Calculate bbox (ensure x1 < x2 and y1 < y2)
        x1 = min(self.rect_start_x, end_x)
        y1 = min(self.rect_start_y, end_y)
        x2 = max(self.rect_start_x, end_x)
        y2 = max(self.rect_start_y, end_y)
        
        # Minimum rectangle size
        if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
            self.drawing = False
            return
        
        bbox = (x1, y1, x2, y2)
        x = x1
        y = y1
        
        # Ask for Tag Details
        tag_name = simpledialog.askstring(
            "Tag Details", 
            "Enter Tag Details (e.g., 'TB1', 'R5', 'F1'):",
            parent=self.root
        )
        
        if not tag_name:
            self.drawing = False
            return  # User cancelled
        
        # Create context menu
        menu = Menu(self.root, tearoff=0)
        
        for category, errors in self.error_categories.items():
            cat_menu = Menu(menu, tearoff=0)
            for error_name, error_template in errors.items():
                cat_menu.add_command(
                    label=error_name,
                    command=lambda en=error_name, et=error_template, c=category, cx=x, cy=y, tn=tag_name, bb=bbox: self.log_error(c, en, et, cx, cy, tn, bb)
                )
            menu.add_cascade(label=f"🔧 {category}", menu=cat_menu)
        
        # Show menu at cursor position
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        
        self.drawing = False
    
    def on_ctrl_click(self, event):
        """Handle Ctrl+Click to select an annotation"""
        if not self.pdf_document:
            return
        
        # Get canvas coordinates
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        
        # Find annotation at this point
        for ann in self.annotations:
            if ann['page'] == self.current_page and 'bbox' in ann:
                bbox = ann['bbox']
                x1, y1, x2, y2 = bbox
                
                # Check if click is inside this annotation
                if ann['type'] == 'ok':
                    # For circles, check distance from center
                    center_x = (x1 + x2) / 2
                    center_y = (y1 + y2) / 2
                    radius = max(abs(x2 - x1), abs(y2 - y1)) / 2
                    distance = ((x - center_x)**2 + (y - center_y)**2)**0.5
                    is_inside = distance <= radius
                else:
                    # For rectangles, check bbox
                    is_inside = x1 <= x <= x2 and y1 <= y <= y2
                
                if is_inside:
                    # Toggle selection
                    if self.selected_annotation is ann:
                        self.selected_annotation = None
                    else:
                        self.selected_annotation = ann
                    self.display_page()
                    return
        
        # If no annotation was clicked, deselect
        self.selected_annotation = None
        self.display_page()
    
    def delete_selected_annotation(self, event=None):
        """Delete the currently selected annotation"""
        if not self.selected_annotation:
            messagebox.showinfo("No Selection", "Please select an annotation first by Ctrl+Clicking on it")
            return
        
        # Confirm deletion
        if messagebox.askyesno("Delete Annotation", "Are you sure you want to delete this annotation?"):
            self.annotations.remove(self.selected_annotation)
            self.selected_annotation = None
            self.display_page()
            messagebox.showinfo("Deleted", "Annotation removed successfully")
    
    def log_error(self, component_type, error_name, error_template, x, y, tag_name, bbox):
        """Log an error to Excel and add annotation"""
        try:
            from openpyxl.styles import Alignment, Border, Side
            
            # Handle special case for short link missing (needs terminals input)
            if error_name == 'Short link missing':
                terminals = simpledialog.askstring(
                    "Terminals", 
                    "Enter terminals where shortlink is expected (e.g., '1-2', '5-6-7'):",
                    parent=self.root
                )
                if not terminals:
                    return  # User cancelled
                punch_text = error_template.format(tag=tag_name, terminals=terminals)
            else:
                # Format the punch text with tag name
                punch_text = error_template.format(tag=tag_name)
            
            # Add annotation
            self.annotations.append({
                'type': 'error',
                'page': self.current_page,
                'x': x / (2.0 * self.zoom_level),
                'y': y / (2.0 * self.zoom_level),
                'bbox': bbox,
                'component': component_type,
                'tag_name': tag_name,
                'error': error_name,
                'punch_text': punch_text,
                'timestamp': datetime.now().isoformat()
            })
            
            # Log to Excel
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Find next empty row (starting from row 11)
            row_num = 11
            while ws[f'E{row_num}'].value is not None:
                row_num += 1
            
            # Add data - with Sr No. but NO Reference No.
            ws[f'C{row_num}'] = self.current_sr_no  # Sr No.
            # D column (Reference No.) left empty
            ws[f'E{row_num}'] = punch_text  # Punch / Action Point
            ws[f'F{row_num}'] = component_type  # Category
            ws[f'G{row_num}'] = os.getlogin()  # Checked By - Name
            ws[f'H{row_num}'] = datetime.now().strftime("%Y-%m-%d")  # Checked By - Date
            
            # Apply formatting
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                cell = ws[f'{col}{row_num}']
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Increment Sr No for next entry
            self.current_sr_no += 1
            
            wb.save(self.excel_file)
            
            self.display_page()
            
            # Show confirmation
            self.root.after(100, lambda: messagebox.showinfo(
                "Logged", 
                f"Error logged: {punch_text}"
            ))
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log error: {str(e)}")
    
    def prev_page(self):
        """Go to previous page"""
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.display_page()
    
    def next_page(self):
        """Go to next page"""
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display_page()
    
    def zoom_in(self):
        """Increase zoom level"""
        if self.zoom_level < 3.0:
            self.zoom_level += 0.25
            self.display_page()
    
    def zoom_out(self):
        """Decrease zoom level"""
        if self.zoom_level > 0.5:
            self.zoom_level -= 0.25
            self.display_page()
    
    def export_annotated_pdf(self):
        """Export PDF with annotations drawn on it"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return
        
        try:
            # Ask for save location
            save_path = filedialog.asksaveasfilename(
                title="Save Annotated PDF",
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"{self.cabinet_id}_annotated.pdf" if self.cabinet_id else "annotated.pdf"
            )
            
            if not save_path:
                return  # User cancelled
            
            # Create a new PDF document
            doc = fitz.open()
            
            # Copy all pages from original and add annotations
            for page_num in range(len(self.pdf_document)):
                # Get the page
                page = self.pdf_document[page_num]
                
                # Copy page to new document
                doc.insert_pdf(self.pdf_document, from_page=page_num, to_page=page_num)
                new_page = doc[page_num]
                
                # Draw annotations for this page
                for ann in self.annotations:
                    if ann['page'] == page_num and 'bbox' in ann:
                        bbox = ann['bbox']
                        # Convert bbox to fitz.Rect (scale back from display coordinates)
                        x1, y1, x2, y2 = bbox
                        x1 = x1 / (2.0 * self.zoom_level)
                        y1 = y1 / (2.0 * self.zoom_level)
                        x2 = x2 / (2.0 * self.zoom_level)
                        y2 = y2 / (2.0 * self.zoom_level)
                        
                        rect = fitz.Rect(x1, y1, x2, y2)
                        
                        if ann['type'] == 'ok':
                            # Green rectangle
                            new_page.draw_rect(rect, color=(0, 1, 0), width=2)
                        else:
                            # Yellow rectangle
                            new_page.draw_rect(rect, color=(1, 1, 0), width=2)
            
            # Save the annotated PDF
            doc.save(save_path)
            doc.close()
            
            messagebox.showinfo("Success", f"Annotated PDF saved to:\n{save_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export annotated PDF: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def open_excel(self):
        """Open the Excel log file"""
        try:
            os.startfile(self.excel_file)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel: {str(e)}")


def main():
    root = tk.Tk()
    app = CircuitInspector(root)
    root.mainloop()


if __name__ == "__main__":
    main()
